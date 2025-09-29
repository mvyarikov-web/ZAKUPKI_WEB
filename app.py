import os
import shutil
import logging
import time
from logging.handlers import TimedRotatingFileHandler
from flask import Flask, render_template, request, jsonify, redirect, url_for, g
from werkzeug.utils import secure_filename
import docx
import openpyxl
import pdfplumber
import re
from flask import send_from_directory
import json
from datetime import datetime
from urllib.parse import unquote
import unicodedata
import docx2txt
import subprocess
from document_processor import DocumentProcessor

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['INDEX_FOLDER'] = 'index'
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max file size
app.config['SEARCH_RESULTS_FILE'] = 'uploads/search_results.json'
app.config['JSON_AS_ASCII'] = False  # корректная кириллица в JSON

# Логирование: файл logs/app.log, ежедневная ротация, хранить 7 архивов
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOGS_DIR = os.path.join(BASE_DIR, 'logs')
os.makedirs(LOGS_DIR, exist_ok=True)
LOG_FILE = os.path.join(LOGS_DIR, 'app.log')

_handler = TimedRotatingFileHandler(
    LOG_FILE, when='midnight', interval=1, backupCount=7, encoding='utf-8'
)
_formatter = logging.Formatter(
    '%(asctime)s %(levelname)s [%(name)s] %(module)s:%(lineno)d - %(message)s'
)
_handler.setFormatter(_formatter)
_handler.setLevel(logging.INFO)

logger = logging.getLogger('zakupki_web')
logger.setLevel(logging.INFO)
logger.addHandler(_handler)
logger.propagate = False

# Настроим логгер Flask
app.logger.handlers = []
app.logger.addHandler(_handler)
app.logger.setLevel(logging.INFO)
app.logger.propagate = False
logging.getLogger('werkzeug').setLevel(logging.WARNING)

# Корневой логгер: подключаем тот же обработчик, чтобы все модули писали в общий лог
root_logger = logging.getLogger()
if _handler not in root_logger.handlers:
    root_logger.addHandler(_handler)
if root_logger.level > logging.INFO:
    root_logger.setLevel(logging.INFO)

# Разрешенные расширения файлов
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'txt'}

# Статусы обработки файлов
file_status = {}  # filename: {'status': 'not_checked'|'processing'|'contains_keywords'|'no_keywords'|'error', 'result': {...}}


def _parse_index_char_counts(index_path: str) -> dict:
    """Парсит _search_index.txt и возвращает {relative_path: char_count} только для файлов из ФС (без zip://, rar://).
    Безопасен к ошибкам формата; игнорирует записи без чисел.
    """
    mapping: dict[str, int] = {}
    if not os.path.exists(index_path):
        return mapping
    try:
        current_title = None
        with open(index_path, 'r', encoding='utf-8', errors='ignore') as f:
            for raw in f:
                line = raw.strip()
                if line.startswith('ЗАГОЛОВОК:'):
                    title = line.split(':', 1)[1].strip()
                    # интересуют только реальные файлы из uploads (без схем)
                    if title and '://' not in title:
                        current_title = title
                    else:
                        current_title = None
                elif current_title and line.startswith('Формат:') and 'Символов:' in line:
                    try:
                        # ожидание шаблона: Формат: ... | Символов: N | ...
                        parts = [p.strip() for p in line.split('|')]
                        for p in parts:
                            if p.startswith('Символов:'):
                                n_str = p.split(':', 1)[1].strip()
                                n = int(''.join(ch for ch in n_str if ch.isdigit())) if n_str else 0
                                mapping[current_title] = n
                                break
                    except Exception:
                        # пропускаем некорректные строки
                        pass
                elif line.startswith('====='):
                    # разделитель — сбрасываем состояние
                    current_title = None
    except Exception:
        app.logger.exception('Ошибка парсинга индекса для char_count')
    return mapping


def _is_safe_subpath(base_dir: str, user_path: str) -> bool:
    """Проверяет, что путь user_path находится внутри base_dir (без обхода через ..).
    Возвращает True, если безопасно.
    """
    try:
        base_abs = os.path.realpath(os.path.abspath(base_dir))
        target_abs = os.path.realpath(os.path.abspath(os.path.join(base_dir, user_path)))
        return os.path.commonpath([base_abs]) == os.path.commonpath([base_abs, target_abs])
    except Exception:
        return False

def safe_filename(filename):
    """Создает безопасное имя файла, сохраняя кириллицу"""
    # Удаляем опасные символы, но сохраняем кириллицу
    import string
    import unicodedata
    
    # Разделяем имя файла и расширение
    name, ext = os.path.splitext(filename)
    
    # Заменяем опасные символы
    dangerous_chars = '<>:"/\\|?*'
    for char in dangerous_chars:
        name = name.replace(char, '_')
    
    # Убираем лишние пробелы и точки
    name = name.strip('. ')
    
    # Если имя стало пустым, даем базовое имя
    if not name:
        name = 'file'
    
    return name + ext

def allowed_file(filename):
    """Проверяет поддержку расширения и исключает временные файлы Office (~$*, $*)."""
    if not filename:
        return False
    base = os.path.basename(filename)
    if base.startswith('~$') or base.startswith('$'):
        return False
    return '.' in base and base.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_search_results():
    """Сохранение результатов поиска в JSON файл"""
    try:
        data = {
            'last_updated': datetime.now().isoformat(),
            'file_status': file_status,
            'last_search_terms': getattr(save_search_results, 'last_terms', ''),
        }
        
        os.makedirs(os.path.dirname(app.config['SEARCH_RESULTS_FILE']), exist_ok=True)
        with open(app.config['SEARCH_RESULTS_FILE'], 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        app.logger.info(f"Результаты сохранены в {app.config['SEARCH_RESULTS_FILE']}")
    except Exception as e:
        app.logger.exception(f"Ошибка сохранения результатов: {str(e)}")

def load_search_results():
    """Загрузка результатов поиска из JSON файла"""
    global file_status
    try:
        if os.path.exists(app.config['SEARCH_RESULTS_FILE']):
            with open(app.config['SEARCH_RESULTS_FILE'], 'r', encoding='utf-8') as f:
                data = json.load(f)
                file_status = data.get('file_status', {})
                save_search_results.last_terms = data.get('last_search_terms', '')
                app.logger.info(f"Результаты загружены из {app.config['SEARCH_RESULTS_FILE']}")
                app.logger.info(f"Загружено статусов: {len(file_status)}")
                return data.get('last_search_terms', '')
    except Exception as e:
        app.logger.exception(f"Ошибка загрузки результатов: {str(e)}")
        file_status = {}
    return ''

def _index_file_path() -> str:
    """Путь к файлу индекса внутри папки index."""
    return os.path.join(app.config['INDEX_FOLDER'], '_search_index.txt')

def clear_search_results():
    """Очистка файла с результатами поиска"""
    global file_status
    try:
        if os.path.exists(app.config['SEARCH_RESULTS_FILE']):
            os.remove(app.config['SEARCH_RESULTS_FILE'])
        file_status = {}
        app.logger.info("Результаты поиска очищены")
        return True
    except Exception as e:
        app.logger.exception(f"Ошибка очистки результатов: {str(e)}")
        return False

def extract_text_from_pdf(file_path):
    """Извлечение текста из PDF файла"""
    text = ""
    try:
        # Попробуем pdfplumber сначала (лучше для сложных PDF)
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
    except:
        try:
            # Если pdfplumber не работает, попробуем pypdf
            import pypdf  # type: ignore
            with open(file_path, 'rb') as file:
                pdf_reader = pypdf.PdfReader(file)
                for page in pdf_reader.pages:
                    page_text = page.extract_text() or ""
                    if page_text:
                        text += page_text + "\n"
        except Exception:
            text = "Ошибка извлечения текста из PDF"
    
    return text

def extract_text_from_docx(file_path):
    """Извлечение текста из DOCX файла"""
    try:
        doc = docx.Document(file_path)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text
    except Exception as e:
        print(f"Ошибка при чтении DOCX файла {file_path}: {str(e)}")
        return "Ошибка извлечения текста из DOCX"

def extract_text_from_doc(file_path):
    """Извлечение текста из DOC файла (старый формат)"""
    try:
        # Читаем файл как бинарный и извлекаем читаемый текст
        with open(file_path, 'rb') as file:
            content = file.read()
            
        # Извлекаем текст из бинарных данных DOC файла
        text = ''
        i = 0
        while i < len(content):
            byte = content[i]
            # Ищем читаемые символы (ASCII и кириллица)
            if 32 <= byte <= 126:  # Обычные ASCII символы
                text += chr(byte)
            elif byte == 0:  # Null байт - пропускаем
                text += ' '
            elif byte in [9, 10, 13]:  # Tab, LF, CR
                text += chr(byte)
            elif byte >= 192 and byte <= 255:  # Возможные символы кириллицы в Windows-1251
                # Пытаемся декодировать как Windows-1251
                try:
                    if i + 1 < len(content):
                        two_bytes = content[i:i+2]
                        decoded = two_bytes.decode('windows-1251', errors='ignore')
                        if decoded and decoded.isprintable():
                            text += decoded
                            i += 1  # Пропускаем следующий байт
                        else:
                            text += ' '
                    else:
                        text += ' '
                except:
                    text += ' '
            else:
                text += ' '
            i += 1
        
        # Очищаем текст от лишних символов и пробелов
        import re
        # Убираем управляющие символы кроме пробелов и переносов строк
        text = re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F-\x9F]', ' ', text)
        # Заменяем множественные пробелы на одиночные
        text = re.sub(r'\s+', ' ', text)
        
        # Извлекаем осмысленные слова (длиннее 2 символов, содержащие буквы)
        words = text.split()
        meaningful_words = []
        for word in words:
            # Очищаем слово от лишних символов
            clean_word = re.sub(r'[^\w\-]', '', word)
            if len(clean_word) > 2 and re.search(r'[a-zA-Zа-яА-ЯёЁ]', clean_word):
                meaningful_words.append(clean_word)
        
        result_text = ' '.join(meaningful_words)
        
        if len(result_text) < 50:  # Если извлеклось мало текста, попробуем docx2txt
            try:
                docx_text = docx2txt.process(file_path)
                if docx_text and len(docx_text.strip()) > len(result_text):
                    result_text = docx_text.strip()
            except:
                pass  # Игнорируем ошибки docx2txt
        
        return result_text if result_text else "Не удалось извлечь достаточно текста из DOC файла"
        
    except Exception as e:
        print(f"Ошибка при чтении DOC файла {file_path}: {str(e)}")
        return "Ошибка извлечения текста из DOC файла. Попробуйте конвертировать файл в DOCX."

def extract_text_from_xlsx(file_path):
    """Извлечение текста из XLSX файла"""
    try:
        workbook = openpyxl.load_workbook(file_path)
        text = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"Лист: {sheet_name}\n"
            for row in sheet.iter_rows():
                row_text = []
                for cell in row:
                    if cell.value is not None:
                        row_text.append(str(cell.value))
                if row_text:
                    text += " | ".join(row_text) + "\n"
        return text
    except Exception as e:
        print(f"Ошибка при чтении XLSX файла {file_path}: {str(e)}")
        return "Ошибка извлечения текста из XLSX"

def extract_text_from_xls(file_path):
    """Извлечение текста из XLS файла (старый формат)"""
    # openpyxl не поддерживает .xls файлы (только .xlsx)
    return "Формат .xls не поддерживается. Пожалуйста, конвертируйте файл в .xlsx"

def extract_text_from_txt(file_path):
    """Извлечение текста из TXT файла"""
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            return file.read()
    except UnicodeDecodeError:
        try:
            with open(file_path, 'r', encoding='cp1251') as file:
                return file.read()
        except:
            try:
                with open(file_path, 'r', encoding='latin-1') as file:
                    return file.read()
            except:
                return "Ошибка извлечения текста из TXT"
    except:
        return "Ошибка извлечения текста из TXT"

def extract_text_from_file(file_path):
    """Универсальная функция для извлечения текста из различных форматов"""
    file_extension = file_path.rsplit('.', 1)[1].lower()
    
    try:
        if file_extension == 'pdf':
            return extract_text_from_pdf(file_path)
        elif file_extension == 'docx':
            return extract_text_from_docx(file_path)
        elif file_extension == 'doc':
            return extract_text_from_doc(file_path)
        elif file_extension == 'xlsx':
            return extract_text_from_xlsx(file_path)
        elif file_extension == 'xls':
            return extract_text_from_xls(file_path)
        elif file_extension == 'txt':
            return extract_text_from_txt(file_path)
        else:
            return "Неподдерживаемый формат файла"
    except Exception as e:
        print(f"Общая ошибка при обработке файла {file_path}: {str(e)}")
        return f"Ошибка обработки файла: {str(e)}"

def search_in_files(search_terms):
    """НОВАЯ логика: поиск по сводному индексу (_search_index.txt), игнорируя заголовки.
    Группирует результаты по файлам и обновляет статусы file_status.
    """
    results = []
    terms = [t.strip() for t in search_terms.split(',') if t.strip()]
    if not terms:
        return results

    uploads = app.config['UPLOAD_FOLDER']
    if not os.path.exists(uploads):
        app.logger.warning('Папка uploads не существует при поиске')
        return results

    # Собираем список реальных файлов для отображения статусов
    files_to_search = []
    for root, dirs, files in os.walk(uploads):
        for fname in files:
            # Исключаем служебный индекс и временные Office-файлы
            if fname == '_search_index.txt' or fname.startswith('~$') or fname.startswith('$'):
                continue
            if allowed_file(fname):
                rel_path = os.path.relpath(os.path.join(root, fname), uploads)
                files_to_search.append(rel_path)

    # Убедимся, что индекс существует (если нет — создадим)
    dp = DocumentProcessor()
    index_path = _index_file_path()
    if not os.path.exists(index_path):
        try:
            dp.create_search_index(uploads)
            app.logger.info('Индекс создан автоматически для поиска')
        except Exception as e:
            app.logger.exception(f'Ошибка создания индекса: {e}')
            return results

    # Поиск по индексу с привязкой к файлам
    try:
        from document_processor.search.searcher import Searcher  # type: ignore
        s = Searcher()
        app.logger.info(f"Поиск по индексу: terms={terms}")
        matches = s.search(index_path, terms, context=80)
    except Exception as e:
        app.logger.exception(f'Ошибка поиска по индексу: {e}')
        matches = []

    # Сгруппируем результаты по файлам (используем key 'title' если появится; fallback — не группировать)
    grouped = {}
    for m in matches:
        title = m.get('title') or m.get('source') or 'индекс'
        d = grouped.setdefault(title, {'found_terms': set(), 'snippets': []})
        d['found_terms'].add(m.get('keyword', ''))
        if len(d['snippets']) < 3:
            d['snippets'].append(m.get('snippet', ''))

    # Обновим статусы известных файлов и подготовим выдачу
    found_files = set()
    for rel_path, data in grouped.items():
        # Статус обновляем только для реальных файлов из uploads
        if rel_path in files_to_search:
            file_status[rel_path] = {
                'status': 'contains_keywords',
                'found_terms': sorted([t for t in data['found_terms'] if t]),
                'context': data['snippets'],
                'processed_at': datetime.now().isoformat()
            }
            found_files.add(rel_path)
        # Добавляем в выдачу, даже если файл виртуальный (например, внутри архива)
        results.append({
            'filename': os.path.basename(rel_path) if isinstance(rel_path, str) else str(rel_path),
            'source': rel_path,
            'path': rel_path if rel_path in files_to_search else None,
            'found_terms': sorted([t for t in data['found_terms'] if t]),
            'context': data['snippets']
        })

    # Реальным файлам без совпадений — статус "нет ключевых слов"
    for rel_path in files_to_search:
        if rel_path not in found_files:
            file_status[rel_path] = {
                'status': 'no_keywords',
                'processed_at': datetime.now().isoformat()
            }

    app.logger.info(f"Поиск завершён: найдено {len(matches)} совпадений, групп: {len(grouped)}")
    return results

@app.route('/')
def index():
    """Главная страница"""
    # Загружаем сохраненные результаты поиска
    last_search_terms = load_search_results()
    
    # Получаем список загруженных файлов с группировкой по папкам
    files_by_folder = {}
    total_files = 0
    
    app.logger.info(f"Проверяем папку: {app.config['UPLOAD_FOLDER']}")
    
    if os.path.exists(app.config['UPLOAD_FOLDER']):
        # Рекурсивно обходим все файлы и папки
        for root, dirs, files in os.walk(app.config['UPLOAD_FOLDER']):
            for filename in files:
                # Скрываем служебный индексный файл и временные файлы Office
                if filename == '_search_index.txt' or filename.startswith('~$') or filename.startswith('$'):
                    continue
                if allowed_file(filename):
                    file_path = os.path.join(root, filename)
                    # Определяем относительную папку
                    relative_folder = os.path.relpath(root, app.config['UPLOAD_FOLDER'])
                    
                    # Правильно определяем название папки
                    if relative_folder == '.':
                        # Файлы в корневой папке uploads
                        folder_display_name = '📁 Загруженные файлы'
                        folder_key = 'root'
                    else:
                        # Файлы в подпапках - берем последнюю часть пути и восстанавливаем оригинальное имя
                        folder_parts = relative_folder.split(os.sep)
                        original_folder_name = folder_parts[-1]
                        # Показываем оригинальное имя папки
                        folder_display_name = f'📂 {original_folder_name}'
                        folder_key = relative_folder
                    
                    file_size = os.path.getsize(file_path)
                    # Получаем статус файла из новой структуры данных
                    file_key = os.path.join(relative_folder, filename) if relative_folder != '.' else filename
                    file_data = file_status.get(file_key, {})
                    status = file_data.get('status', 'not_checked')
                    
                    # Получаем оригинальное имя файла если оно сохранено, иначе используем текущее
                    display_name = file_data.get('original_name', filename)
                    
                    file_info = {
                        'name': display_name,
                        'size': file_size,
                        'status': status,
                        'path': file_key,
                        'relative_folder': relative_folder
                    }
                    
                    if folder_key not in files_by_folder:
                        files_by_folder[folder_key] = {
                            'display_name': folder_display_name,
                            'relative_path': relative_folder,
                            'files': []
                        }
                    
                    files_by_folder[folder_key]['files'].append(file_info)
                    total_files += 1
                    
                    app.logger.debug(f"Добавлен файл: {filename} в папку {folder_display_name}, размер: {file_size}, статус: {status}")
    else:
        app.logger.warning("Папка uploads не существует")
    
    app.logger.info(f"Всего файлов для отображения: {total_files}, папок: {len(files_by_folder)}")
    return render_template(
        'index.html',
        files_by_folder=files_by_folder,
        total_files=total_files,
        last_search_terms=last_search_terms,
        file_status=file_status,
    )

@app.route('/upload', methods=['POST'])
def upload_files():
    """Обработка загрузки файлов"""
    if 'files' not in request.files:
        return jsonify({'error': 'Файлы не выбраны'}), 400
    
    files = request.files.getlist('files')
    uploaded_files = []
    
    for file in files:
        if file and file.filename != '':
            original_filename = file.filename
            app.logger.info(f"Загружается файл: {original_filename}")
            # Пропускаем временные файлы Office (например, ~$file.docx)
            base_name = os.path.basename(original_filename)
            if base_name.startswith('~$') or base_name.startswith('$'):
                app.logger.info(f"Пропуск временного файла Office: {original_filename}")
                continue
            
            if allowed_file(original_filename):
                # Обрабатываем путь из webkitRelativePath если есть (для папок)
                relative_path = request.form.get('webkitRelativePath', '')
                if not relative_path:
                    # Пытаемся получить путь из самого файла (если поддерживается браузером)
                    relative_path = getattr(file, 'filename', original_filename)
                
                # Разделяем путь на папки и имя файла
                path_parts = relative_path.split('/') if '/' in relative_path else [relative_path]
                filename = path_parts[-1]  # Последняя часть - имя файла
                folder_parts = path_parts[:-1]  # Все предыдущие части - путь к папке
                
                # Создаем безопасные имена для папок и файла, сохраняя кириллицу
                safe_folder_parts = [safe_filename(part) for part in folder_parts if part]
                safe_filename_result = safe_filename(filename)
                
                # Восстанавливаем расширение если safe_filename его удалил
                if '.' not in safe_filename_result and '.' in filename:
                    extension = filename.rsplit('.', 1)[1].lower()
                    safe_filename_result = safe_filename_result + '.' + extension
                
                # Создаем полный путь к папке
                if safe_folder_parts:
                    target_folder = os.path.join(app.config['UPLOAD_FOLDER'], *safe_folder_parts)
                    os.makedirs(target_folder, exist_ok=True)
                else:
                    target_folder = app.config['UPLOAD_FOLDER']
                
                # Проверяем уникальность имени файла
                counter = 1
                base_name, extension = os.path.splitext(safe_filename_result) if '.' in safe_filename_result else (safe_filename_result, '')
                final_filename = safe_filename_result
                
                while os.path.exists(os.path.join(target_folder, final_filename)):
                    final_filename = f"{base_name}_{counter}{extension}"
                    counter += 1
                
                # Сохраняем файл
                file_path = os.path.join(target_folder, final_filename)
                file.save(file_path)
                
                # Устанавливаем статус "не проверено" для новых файлов
                # Используем относительный путь как ключ для уникальности
                file_key = os.path.join(*safe_folder_parts, final_filename) if safe_folder_parts else final_filename
                file_status[file_key] = {'status': 'not_checked', 'result': None, 'original_name': filename}
                uploaded_files.append(final_filename)
                
                app.logger.info(f"Файл сохранён: {file_path}, ключ: {file_key}")
            else:
                return jsonify({'error': f'Неподдерживаемый тип файла: {original_filename}'}), 400
    
    app.logger.info(f"Загружено файлов: {len(uploaded_files)}")
    return jsonify({'success': True, 'uploaded_files': uploaded_files})

@app.route('/search', methods=['POST'])
def search():
    """Поиск по ключевым словам"""
    search_terms = request.json.get('search_terms', '')
    
    if not search_terms.strip():
        return jsonify({'error': 'Введите ключевые слова для поиска'}), 400
    
    # Валидация: не более 10 терминов, длина 2..64, удаление дубликатов
    raw_terms = [t.strip() for t in search_terms.split(',') if t.strip()]
    if len(raw_terms) > 50:  # жёсткий предел на вход
        raw_terms = raw_terms[:50]
    filtered = []
    seen = set()
    for t in raw_terms[:10]:
        if 2 <= len(t) <= 64 and t.lower() not in seen:
            seen.add(t.lower())
            filtered.append(t)
    if not filtered:
        return jsonify({'error': 'Слишком короткие/длинные или пустые ключевые слова'}), 400

    # Новый поиск через индекс
    app.logger.info(f"Запрос поиска: terms='{','.join(filtered)}' (из {len(raw_terms)} входных)")
    results = search_in_files(','.join(filtered))
    
    # Сохраняем результаты поиска
    save_search_results.last_terms = search_terms
    save_search_results()
    app.logger.info(f"Результаты поиска сохранены. Найдено групп: {len(results)}")
    
    return jsonify({'results': results})

@app.route('/build_index', methods=['POST'])
def build_index():
    """Явная сборка индекса по папке uploads."""
    uploads = app.config['UPLOAD_FOLDER']
    if not os.path.exists(uploads):
        return jsonify({'success': False, 'message': 'Папка uploads не найдена'}), 400
    try:
        dp = DocumentProcessor()
        app.logger.info("Запуск явной сборки индекса для uploads")
        # Создаём индекс в uploads, затем переносим в index/
        tmp_index_path = dp.create_search_index(uploads)
        os.makedirs(app.config['INDEX_FOLDER'], exist_ok=True)
        index_path = _index_file_path()
        try:
            if os.path.exists(tmp_index_path):
                shutil.move(tmp_index_path, index_path)
            else:
                # На всякий случай, если реализация уже пишет в index_folder
                if os.path.exists(index_path):
                    pass
        except Exception:
            app.logger.exception('Не удалось переместить индекс в папку index')
        size = os.path.getsize(index_path) if os.path.exists(index_path) else 0
        app.logger.info(f"Индекс собран: {index_path}, размер: {size} байт")
        # Обновим количество распознанных символов по каждому реальному файлу и статусы ошибок/неподдержки
        try:
            counts = _parse_index_char_counts(index_path)
            # Список всех файлов в uploads
            all_files: list[str] = []
            for root, dirs, files in os.walk(uploads):
                for fname in files:
                    if fname == '_search_index.txt' or fname.startswith('~$') or fname.startswith('$'):
                        continue
                    rel_path = os.path.relpath(os.path.join(root, fname), uploads)
                    all_files.append(rel_path)
            for rel_path in all_files:
                ext_ok = allowed_file(rel_path)
                entry = file_status.get(rel_path, {})
                if not ext_ok:
                    # Неподдерживаемый формат
                    entry.update({'status': 'unsupported', 'error': 'Неподдерживаемый формат', 'char_count': 0, 'processed_at': datetime.now().isoformat()})
                else:
                    cc = counts.get(rel_path)
                    if cc is None:
                        # Поддерживаемый, но нет записи в индексе — ошибка чтения/индексации
                        entry.update({'status': entry.get('status', 'error' if entry.get('status') in (None, 'not_checked') else entry.get('status')),
                                      'error': entry.get('error') or 'Ошибка чтения или не проиндексирован',
                                      'char_count': 0,
                                      'processed_at': datetime.now().isoformat()})
                    else:
                        # Есть счётчик символов — не трогаем статус поиска, только дополняем метрикой
                        entry.update({'char_count': cc, 'processed_at': datetime.now().isoformat()})
                        # если 0 символов, оставим это как индикатор качества (UI подсветит)
                file_status[rel_path] = entry
            # Сохраним
            save_search_results()
        except Exception:
            app.logger.exception('Не удалось обновить char_count по индексу')
        return jsonify({'success': True, 'index_path': index_path, 'size': size})
    except Exception as e:
        app.logger.exception("Ошибка при сборке индекса")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/delete/<path:filepath>', methods=['DELETE'])
def delete_file(filepath):
    """Удаление файла"""
    try:
        # Декодируем URL-кодированный путь к файлу
        from urllib.parse import unquote
        decoded_filepath = unquote(filepath)
        app.logger.info(f"Попытка удалить файл: {decoded_filepath}")
        
        # Проверяем, что это не попытка выйти за пределы папки uploads
        if not _is_safe_subpath(app.config['UPLOAD_FOLDER'], decoded_filepath):
            return jsonify({'error': 'Недопустимый путь к файлу'}), 400
        
        # Полный путь к файлу
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], decoded_filepath)
        app.logger.debug(f"Полный путь к файлу: {file_path}")
        
        if os.path.exists(file_path) and os.path.isfile(file_path):
            os.remove(file_path)
            # Удаляем из статусов файлов все возможные варианты ключей
            file_status.pop(decoded_filepath, None)
            file_status.pop(os.path.basename(decoded_filepath), None)
            
            # Сохраняем обновленные результаты
            save_search_results()
            app.logger.info(f"Файл {decoded_filepath} успешно удалён")
            return jsonify({'success': True})
        else:
            app.logger.warning(f"Файл не найден: {file_path}")
            return jsonify({'error': 'Файл не найден'}), 404
            
    except Exception as e:
        app.logger.exception(f"Ошибка при удалении файла: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/delete_folder/<path:folder_path>', methods=['DELETE'])
def delete_folder(folder_path):
    """Удаление папки со всеми файлами"""
    try:
        from urllib.parse import unquote
        decoded_folder_path = unquote(folder_path)
        app.logger.info(f"Попытка удалить папку: {decoded_folder_path}")
        
        # Проверяем безопасность пути
        if decoded_folder_path != 'root' and not _is_safe_subpath(app.config['UPLOAD_FOLDER'], decoded_folder_path):
            return jsonify({'error': 'Недопустимый путь к папке'}), 400
        
        if decoded_folder_path == 'root':
            # Удаление всех файлов из корневой папки uploads
            target_folder = app.config['UPLOAD_FOLDER']
            deleted_files = []
            
            if os.path.exists(target_folder):
                for item in os.listdir(target_folder):
                    item_path = os.path.join(target_folder, item)
                    if os.path.isfile(item_path):
                        os.remove(item_path)
                        deleted_files.append(item)
                        # Удаляем из статусов
                        file_status.pop(item, None)
                        
            message = f'Удалено {len(deleted_files)} файлов из корневой папки'
        else:
            # Удаление подпапки
            target_folder = os.path.join(app.config['UPLOAD_FOLDER'], decoded_folder_path)
            
            if not os.path.exists(target_folder):
                return jsonify({'error': 'Папка не найдена'}), 404
            
            # Собираем список файлов для удаления из статусов
            files_to_remove_from_status = []
            for root, dirs, files in os.walk(target_folder):
                for filename in files:
                    rel_path = os.path.relpath(os.path.join(root, filename), app.config['UPLOAD_FOLDER'])
                    files_to_remove_from_status.append(rel_path)
            
            # Удаляем папку рекурсивно
            shutil.rmtree(target_folder)
            
            # Удаляем файлы из статусов
            for file_key in files_to_remove_from_status:
                file_status.pop(file_key, None)
                file_status.pop(os.path.basename(file_key), None)
            
            message = f'Папка "{decoded_folder_path}" успешно удалена'
        
        # Сохраняем обновленные результаты
        save_search_results()
        app.logger.info(message)
        
        return jsonify({
            'success': True,
            'message': message
        })
        
    except Exception as e:
        app.logger.exception(f"Ошибка при удалении папки: {str(e)}")
        return jsonify({'error': f'Ошибка удаления папки: {str(e)}'}), 500


@app.get('/download/<path:filepath>')
def download_file(filepath: str):
    """Безопасная выдача файла из uploads."""
    try:
        from urllib.parse import unquote
        decoded = unquote(filepath)
        if not _is_safe_subpath(app.config['UPLOAD_FOLDER'], decoded):
            return jsonify({'error': 'Недопустимый путь'}), 400
        # Запрещаем скачивание скрытых и служебных файлов
        if os.path.basename(decoded).startswith('.'):
            return jsonify({'error': 'Файл недоступен'}), 403
        # Ограничим типы по whitelist
        if not allowed_file(decoded):
            return jsonify({'error': 'Неподдерживаемый тип файла'}), 400
        directory = app.config['UPLOAD_FOLDER']
        # Разбиваем путь на каталог и имя
        full_path = os.path.join(directory, decoded)
        if not os.path.isfile(full_path):
            return jsonify({'error': 'Файл не найден'}), 404
        folder = os.path.dirname(decoded)
        fname = os.path.basename(decoded)
        return send_from_directory(os.path.join(directory, folder) if folder else directory, fname, as_attachment=False)
    except Exception as e:
        app.logger.exception('download_file error')
        return jsonify({'error': str(e)}), 500


@app.get('/files_json')
def files_json():
    """JSON-список файлов в uploads, сгруппированный по папкам (упрощённо)."""
    uploads = app.config['UPLOAD_FOLDER']
    if not os.path.exists(uploads):
        return jsonify({'folders': {}, 'total_files': 0})
    files_by_folder = {}
    total_files = 0
    for root, dirs, files in os.walk(uploads):
        rel_dir = os.path.relpath(root, uploads)
        for filename in files:
            # Скрываем служебный индексный файл и временные файлы Office
            if filename == '_search_index.txt' or filename.startswith('~$') or filename.startswith('$'):
                continue
            if not allowed_file(filename):
                continue
            file_path = os.path.join(root, filename)
            rel_path = os.path.normpath(os.path.join(rel_dir, filename)) if rel_dir != '.' else filename
            folder_key = 'root' if rel_dir == '.' else rel_dir
            folder_name = '📁 Загруженные файлы' if rel_dir == '.' else f'📂 {os.path.basename(rel_dir)}'
            if folder_key not in files_by_folder:
                files_by_folder[folder_key] = {'display_name': folder_name, 'relative_path': rel_dir if rel_dir != '.' else '', 'files': []}
            meta = file_status.get(rel_path, {})
            files_by_folder[folder_key]['files'].append({
                'name': filename,
                'size': os.path.getsize(file_path),
                'status': meta.get('status', 'not_checked'),
                'char_count': meta.get('char_count'),
                'error': meta.get('error'),
                'path': rel_path
            })
            total_files += 1
    return jsonify({'folders': files_by_folder, 'total_files': total_files})


@app.get('/index_status')
def index_status():
    """Статус индексного файла index/_search_index.txt: наличие, размер, mtime, записи."""
    try:
        idx = _index_file_path()
        exists = os.path.exists(idx)
        if not exists:
            return jsonify({'exists': False})
        size = os.path.getsize(idx)
        mtime = datetime.fromtimestamp(os.path.getmtime(idx)).isoformat()
        # Подсчёт числа документов: количество строк-разделителей "====..." делим на 1 (каждая запись начинается с бара)
        try:
            with open(idx, 'r', encoding='utf-8', errors='ignore') as f:
                count = sum(1 for line in f if line.strip().startswith('='))
            # Запись начинается с одной линии бара, но между записями тоже бар, поэтому count/2?
            # Формат: для каждой записи идёт две линии бара (верх и низ), значит число записей приблизительно count/2.
            entries = max(0, count // 2)
        except Exception:
            entries = None
        return jsonify({'exists': True, 'size': size, 'mtime': mtime, 'entries': entries})
    except Exception as e:
        app.logger.exception('index_status error')
        return jsonify({'exists': False, 'error': str(e)}), 500


@app.errorhandler(413)
def request_entity_too_large(e):
    return jsonify({'error': 'Файл слишком большой. Лимит 100MB.'}), 413

# Просмотр сводного файла индекса в отдельной вкладке
@app.get('/view_index')
def view_index():
    idx = _index_file_path()
    if not os.path.exists(idx):
        return jsonify({'error': 'Индекс не найден'}), 404
    try:
        with open(idx, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        from flask import Response
        return Response(content, mimetype='text/plain; charset=utf-8')
    except Exception as e:
        app.logger.exception('Ошибка чтения сводного файла индекса')
        return jsonify({'error': str(e)}), 500

# Логирование HTTP-запросов: время, путь, статус
@app.before_request
def _start_timer():
    g._start_time = time.time()

@app.after_request
def _log_request(response):
    try:
        duration_ms = (time.time() - getattr(g, '_start_time', time.time())) * 1000.0
        app.logger.info(f"{request.remote_addr} {request.method} {request.path} {response.status_code} {duration_ms:.1f}ms")
    except Exception:
        pass
    return response

# Очистка результатов поиска
@app.route('/clear_results', methods=['POST'])
def clear_results():
    """Очистка всех сохраненных результатов поиска"""
    if clear_search_results():
        return jsonify({'success': True, 'message': 'Результаты поиска успешно очищены'})
    else:
        return jsonify({'success': False, 'message': 'Ошибка очистки результатов'})

@app.get('/health')
def health():
    return jsonify({'status': 'ok'}), 200


def _ensure_uploads_dir():
    """Гарантирует наличие папки uploads при импорте/старте."""
    try:
        if not os.path.exists(app.config['UPLOAD_FOLDER']):
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    except Exception:
        app.logger.exception('Не удалось создать папку uploads')


# Выполним инициализацию при импорте (без тяжёлых побочных эффектов)
_ensure_uploads_dir()
if not hasattr(save_search_results, 'last_terms'):
    save_search_results.last_terms = ''


if __name__ == '__main__':
    _ensure_uploads_dir()
    # Инициализируем атрибут для сохранения последних ключевых слов
    save_search_results.last_terms = ''
    # Всегда запускаем на 8081: предварительно освобождаем порт и стартуем там
    import subprocess
    import signal
    def _pids_on_port(p: int) -> list[int]:
        try:
            out = subprocess.check_output(["lsof", f"-ti:{p}"] , text=True)
            return [int(x.strip()) for x in out.splitlines() if x.strip().isdigit()]
        except Exception:
            return []
    def _free_port(p: int) -> int:
        killed = 0
        for pid in _pids_on_port(p):
            try:
                os.kill(pid, signal.SIGKILL)
                killed += 1
            except Exception:
                pass
        return killed

    try:
        killed = _free_port(8081)
        app.logger.info(f"Предзапуск: освобождён порт 8081, завершено процессов = {killed}")
    except Exception:
        app.logger.exception('Не удалось освободить 8081 перед запуском')

    app.logger.info("Запуск Flask на порту 8081 (жёстко)")
    app.run(debug=False, use_reloader=False, host='127.0.0.1', port=8081)