"""Универсальный экстрактор текста (расширенная версия, graceful degrade).

Поддержка:
- TXT (автодетект кодировки)
- JSON, CSV/TSV, XML/HTML (как текст)
- PDF (pdfplumber → pypdf, иначе пустая строка)
- DOCX (python-docx)
- XLSX (openpyxl)
- XLS (xlrd — опционально)

Поддерживает два режима:
1. Чтение из файловой системы (file_path: str)
2. Чтение из памяти (file_bytes: bytes + extension: str)

Все тяжёлые импорты обёрнуты в try/except. При любой ошибке возвращается пустая строка.
"""
from __future__ import annotations

from pathlib import Path
from typing import Optional, Union
from io import BytesIO
import chardet  # type: ignore


def _read_text_with_encoding(data: bytes) -> str:
    """Чтение текста с автодетектом кодировки."""
    guess = chardet.detect(data) or {}
    enc = guess.get('encoding') or 'utf-8'
    for codec in (enc, 'utf-8', 'cp1251', 'cp866', 'latin-1'):
        try:
            return data.decode(codec, errors='ignore')
        except Exception:
            continue
    return ''


def _clean_html_xml(raw: str) -> str:
    """Очистка HTML/XML от тегов и entities."""
    import re
    # Удаляем скрипты/стили
    raw = re.sub(r'<script[\s\S]*?</script>', ' ', raw, flags=re.IGNORECASE)
    raw = re.sub(r'<style[\s\S]*?</style>', ' ', raw, flags=re.IGNORECASE)
    # Удаляем все теги
    raw = re.sub(r'<[^>]+>', ' ', raw)
    # HTML entities упрощённо
    raw = raw.replace('&nbsp;', ' ').replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')
    # Числовые сущности (&#NNNN; и &#xHH;) → символы
    def _decode_entity(m):
        ent = m.group(1)
        try:
            if ent.lower().startswith('x'):
                return chr(int(ent[1:], 16))
            return chr(int(ent))
        except Exception:
            return ' '
    raw = re.sub(r'&#(x?[0-9A-Fa-f]+);', _decode_entity, raw)
    # Сжатие пробелов
    raw = re.sub(r'[ \t]+', ' ', raw)
    raw = re.sub(r'\s*\n\s*', '\n', raw)
    return raw.strip()


def extract_text_from_bytes(file_bytes: bytes, extension: str) -> str:
    """Извлечь текст из бинарных данных файла.

    Args:
        file_bytes: Содержимое файла
        extension: Расширение файла без точки (например, 'pdf', 'docx')

    Returns:
        str: Извлечённый текст или пустая строка
    """
    try:
        ext = extension.lower().lstrip('.')
        
        # Текстовые форматы
        if ext in {'txt', 'json', 'csv', 'tsv', 'xml', 'html', 'htm'}:
            raw = _read_text_with_encoding(file_bytes)
            if ext in {'html', 'htm', 'xml'}:
                raw = _clean_html_xml(raw)
            return raw
        
        # PDF
        if ext == 'pdf':
            try:
                import pdfplumber  # type: ignore
                text_parts = []
                with pdfplumber.open(BytesIO(file_bytes)) as pdf:
                    for page in pdf.pages:
                        t = page.extract_text() or ''
                        if t:
                            text_parts.append(t)
                return '\n'.join(text_parts)
            except Exception:
                # Fallback на pypdf
                try:
                    from pypdf import PdfReader  # type: ignore
                    reader = PdfReader(BytesIO(file_bytes))
                    parts = []
                    for page in reader.pages:
                        t = page.extract_text() or ''
                        if t:
                            parts.append(t)
                    return '\n'.join(parts)
                except Exception:
                    return ''
        
        # DOCX
        if ext == 'docx':
            try:
                from docx import Document  # type: ignore
                doc = Document(BytesIO(file_bytes))
                paras = [para.text for para in doc.paragraphs if para.text]
                return '\n'.join(paras)
            except Exception:
                return ''
        
        # XLSX
        if ext == 'xlsx':
            try:
                from openpyxl import load_workbook  # type: ignore
                wb = load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=True)
                parts = []
                for ws in wb.worksheets:
                    for row in ws.iter_rows(values_only=True):
                        vals = [str(v) for v in row if v is not None]
                        if vals:
                            parts.append('\t'.join(vals))
                return '\n'.join(parts)
            except Exception:
                return ''
        
        # XLS (старый формат)
        if ext == 'xls':
            try:
                import xlrd  # type: ignore
                book = xlrd.open_workbook(file_contents=file_bytes)
                parts = []
                for si in range(book.nsheets):
                    sh = book.sheet_by_index(si)
                    for ri in range(sh.nrows):
                        row = sh.row_values(ri)
                        vals = [str(v) for v in row if v not in (None, '')]
                        if vals:
                            parts.append('\t'.join(vals))
                return '\n'.join(parts)
            except Exception:
                return ''
        
        return ''
    except Exception:
        return ''


def extract_text(file_path: str) -> str:
    """Вернуть извлечённый текст из файла. Ошибки глушатся (graceful degrade).

    Args:
        file_path: Путь к файлу

    Returns:
        str: Текстовое содержимое (или пустая строка)
    """
    try:
        p = Path(file_path)
        if not p.exists() or not p.is_file():
            return ''
        
        # Читаем файл в память и используем общую функцию
        file_bytes = p.read_bytes()
        ext = p.suffix.lower().lstrip('.')
        return extract_text_from_bytes(file_bytes, ext)
    except Exception:
        return ''
